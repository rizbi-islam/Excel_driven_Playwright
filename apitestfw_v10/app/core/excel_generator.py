"""
app/core/excel_generator.py — Generate downloadable Excel test suites.

Flow: user enters base URL + endpoints → system probes → generates Excel
→ user downloads, reviews/edits → uploads to run page.
No DB insert happens here. User stays in control.

Sheets: README | Smoke | Regression | Load | Stress | Security
"""
import io
import json

import openpyxl
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter

COLORS = {
    "smoke":      "1565C0",
    "regression": "00695C",
    "load":       "E65100",
    "stress":     "B71C1C",
    "security":   "4A148C",
    "readme":     "263238",
    "white":      "FFFFFF",
    "alt":        "F8FAFC",
}

COMMON_COLS = [
    ("name",            "Test Name",         30),
    ("method",          "Method",             9),
    ("endpoint",        "Endpoint / Path",   38),
    ("headers",         "Headers (JSON)",    28),
    ("body",            "Body (JSON)",       28),
    ("params",          "Params (JSON)",     28),
    ("expected_status", "Expected Status",   16),
    ("auth_type",       "Auth Type",         13),
    ("auth_token",      "Auth Token",        24),
    ("max_response_ms", "Max MS",            11),
    ("assertions",      "Assertions (JSON)", 32),
    ("tags",            "Tags",              18),
    ("description",     "Description",       36),
]
LOAD_EXTRA   = [("concurrency","Concurrency",13), ("duration_sec","Duration (sec)",15)]
STRESS_EXTRA = [("start_users","Start Users",13), ("peak_users","Peak Users",12), ("ramp_sec","Ramp (sec)",12)]
SEC_EXTRA    = [("check_types","Checks (comma-sep)",38)]

_thin   = Side(border_style="thin", color="DDDDDD")
_border = Border(top=_thin, left=_thin, right=_thin, bottom=_thin)


def _hfill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _afill() -> PatternFill:
    return PatternFill("solid", fgColor=COLORS["alt"])


def _write_sheet(wb, title: str, columns: list, rows: list, color: str):
    ws = wb.create_sheet(title)
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = True

    # Header row
    for ci, (_, label, width) in enumerate(columns, 1):
        cell = ws.cell(1, ci, label)
        cell.font      = Font(bold=True, color=COLORS["white"], size=10)
        cell.fill      = _hfill(color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _border
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 22

    # Data rows
    for ri, row_data in enumerate(rows, 2):
        alt = ri % 2 == 0
        for ci, (key, _, _) in enumerate(columns, 1):
            val = row_data.get(key, "")
            if isinstance(val, (dict, list)):
                val = json.dumps(val, ensure_ascii=False)
            cell = ws.cell(ri, ci, val if val is not None else "")
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border    = _border
            if alt:
                cell.fill = _afill()
        ws.row_dimensions[ri].height = 18


def _readme_sheet(wb, base_url: str, endpoints: list):
    ws = wb.create_sheet("README", 0)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 80
    ws.sheet_view.showGridLines = False

    h_fill = _hfill(COLORS["readme"])
    h_font = Font(bold=True, color=COLORS["white"], size=13)
    s_font = Font(bold=True, color=COLORS["readme"], size=10)
    n_font = Font(size=10)
    g_font = Font(size=9, color="888888")

    data = [
        ("API Test Framework v10 — Auto-Generated Test Suite", None, "title"),
        ("", None, None),
        ("Base URL:", base_url, "sub"),
        ("", None, None),
        ("HOW TO USE", None, "sub"),
        ("1.", "Review each sheet: Smoke, Regression, Load, Stress, Security.", None),
        ("2.", "Edit rows: add/remove cases, adjust expected_status, body, params.", None),
        ("3.", "JSON cells: paste valid JSON objects {} or arrays [].", None),
        ("4.", "Leave cells empty when API has no body or no params — that is fine.", None),
        ("5.", "Upload the file on the matching Run page (e.g. Smoke → /run/smoke).", None),
        ("6.", "After run completes, click 'Save to Library' to reuse cases.", None),
        ("", None, None),
        ("ENDPOINTS ANALYZED", None, "sub"),
    ]
    for ep in endpoints:
        data.append((
            f"  {ep.get('method','GET')}",
            f"{ep.get('path','')}   {ep.get('name','') or ''}",
            "ep"
        ))

    for ri, (a, b, style) in enumerate(data, 1):
        ca = ws.cell(ri, 1, a)
        if style == "title":
            ca.font = h_font; ca.fill = h_fill
            ws.row_dimensions[ri].height = 26
        elif style == "sub":
            ca.font = s_font
        elif style == "ep":
            ca.font = g_font
        else:
            ca.font = n_font
        if b is not None:
            cb = ws.cell(ri, 2, b)
            cb.font      = g_font if style == "ep" else n_font
            cb.alignment = Alignment(vertical="center")


class ExcelGenerator:

    @staticmethod
    def generate(base_url: str, endpoints: list, analyses: list) -> bytes:
        """
        endpoints : [{"method":"GET","path":"/users","name":"...", "auth_type":"none", ...}]
        analyses  : [{"status":200,"body_sample":{},"error":None}] — one per endpoint
        Returns   : bytes of the .xlsx file
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        _readme_sheet(wb, base_url, endpoints)

        smoke_rows = regression_rows = load_rows = stress_rows = security_rows = []
        smoke_rows = []
        regression_rows = []
        load_rows = []
        stress_rows = []
        security_rows = []

        for ep, an in zip(endpoints, analyses):
            method    = (ep.get("method") or "GET").upper()
            path      = ep.get("path", "/")
            name      = ep.get("name") or path.strip("/").replace("/", " ").title() or "Test"
            st        = an.get("status", 200)
            body_s    = an.get("body_sample")
            has_body  = method in ("POST", "PUT", "PATCH")
            sample_b  = json.dumps(body_s, ensure_ascii=False) if body_s and has_body else ""
            auth_type = ep.get("auth_type", "none")
            auth_tok  = ep.get("auth_token", "")

            base = dict(
                name=name, method=method, endpoint=path,
                headers="", body=sample_b, params="",
                expected_status=st,
                auth_type=auth_type, auth_token=auth_tok,
                max_response_ms=5000, assertions="",
                tags=method.lower(), description=ep.get("description",""),
            )

            smoke_rows.append({**base, "name": f"[Smoke] {name}", "max_response_ms": 3000})

            regression_rows.append({**base, "name": f"[Regression] {name} — Happy Path"})
            regression_rows.append({
                **base, "name": f"[Regression] {name} — Not Found",
                "endpoint": path.rstrip("/") + "/999999",
                "expected_status": 404,
                "description": "Expect 404 for non-existent resource.",
            })
            if auth_type != "none":
                regression_rows.append({
                    **base, "name": f"[Regression] {name} — Unauthorized",
                    "auth_type": "none", "auth_token": "",
                    "expected_status": 401,
                    "description": "Expect 401 without auth.",
                })

            load_rows.append({
                **base, "name": f"[Load] {name}",
                "concurrency": 10, "duration_sec": 30, "max_response_ms": 2000,
            })

            stress_rows.append({
                **base, "name": f"[Stress] {name}",
                "start_users": 1, "peak_users": 50, "ramp_sec": 60,
            })

            security_rows.append({
                **base, "name": f"[Security] {name}",
                "check_types": "sqli,xss,cors,auth_bypass,https,info_disclosure,rate_limit",
            })

        _write_sheet(wb, "Smoke",      COMMON_COLS,                        smoke_rows,      COLORS["smoke"])
        _write_sheet(wb, "Regression", COMMON_COLS,                        regression_rows, COLORS["regression"])
        _write_sheet(wb, "Load",       COMMON_COLS + LOAD_EXTRA,           load_rows,       COLORS["load"])
        _write_sheet(wb, "Stress",     COMMON_COLS + STRESS_EXTRA,         stress_rows,     COLORS["stress"])
        _write_sheet(wb, "Security",   COMMON_COLS[:6] + SEC_EXTRA + COMMON_COLS[9:],
                     security_rows, COLORS["security"])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()
